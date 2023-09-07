from flask import Flask, render_template, request, redirect
import openpyxl
from openpyxl import Workbook
import pathlib


app = Flask(__name__)
app.config['SECRET_KEY'] = 'root'
arquivo = pathlib.Path("Clientes.xlsx")
if arquivo.exists():
    pass
else:
    arquivo = Workbook()
    folha = arquivo.active
    folha['A1'] = "Nome"
    folha['B1'] = "Whastapp"
    folha['C1'] = "Email"
    folha['D1'] = "Autoriza_contato"
    folha['E1'] = "categoria"
    arquivo.save("Clientes.xlsx")


@app.route('/')
def home():
    return render_template('index.html')


@app.route('/cad', methods=['GET', 'POST'])
def cad():
    arquivo = openpyxl.load_workbook("Clientes.xlsx")
    folha = arquivo.active
    nome = request.form.get("nome")
    whastapp = request.form.get("whatsapp")
    email = request.form.get("email")
    categoria = request.form.get("tipo_contato")
    autoriza_contato = request.form.get("autoriza_contato")
    if request.form.get('nome') != "" and request.form.get('email') != "" and request.form.get('whastapp') != "" and request.form.get('autoriza_contato') != "":
        return redirect('/')
    else:
        folha.cell(column=1, row=folha.max_row+1, value=nome)
        folha.cell(column=2, row=folha.max_row,   value=whastapp)
        folha.cell(column=3, row=folha.max_row,   value=email)
        folha.cell(column=4, row=folha.max_row,   value=autoriza_contato)
        folha.cell(column=5, row=folha.max_row,   value=categoria)
        arquivo.save(r"Clientes.xlsx")
        print("---------------Dados Salvos com Sucesso!-----------------")
        return redirect('/')


if __name__ in "__main__":
    app.run(debug=True)
