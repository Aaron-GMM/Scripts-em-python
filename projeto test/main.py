from flask import Flask, render_template, redirect, request
import openpyxl
from openpyxl import Workbook
import pathlib
import pandas as pd
app = Flask(__name__)
app.config['SECRET_KEY'] = 'root'
arquivo = pathlib.Path("clientes.xlsx")
if arquivo.exists():
    pass
else:
    arquivo = Workbook()
    folha = arquivo.active
    folha['A1'] = "Email:"
    folha['B1'] = "Senha:"
    arquivo.save("clientes.xlsx")


@app.route('/')
def home():
    return render_template('index.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    arquivo = openpyxl.load_workbook('clientes.xlsx')
    folha = arquivo.active
    email = request.form.get('email')
    senha = request.form.get('senha')
    planilha = pd.read_excel('clientes.xlsx')
    resultado = planilha.loc[planilha['Email:'] == email, 'Email:']
    print(resultado)
# Verificar se algum email foi encontrado
    if not resultado.empty:
        print(f'O email  é: {resultado.iloc[0]}')
    
        return render_template('index.html')
    else:
        print(f'Cliente não encontrado.')
        folha.cell(column=1, row=folha.max_row+1, value=email)
        folha.cell(column=2, row=folha.max_row, value=senha)
        arquivo.save(r"clientes.xlsx")
        print("dados salvos com sucesso")
        return redirect('/')


if __name__ in "__main__":
    app.run(debug=True)
