from flask import Flask, render_template, redirect, request
import openpyxl
from openpyxl import Workbook
import pathlib
import pandas as pd
class Planinha():
    def criar_planilha():
        arquivo = pathlib.Path("clientes.xlsx")
        if arquivo.exists():
            pass
        else:
            arquivo = Workbook()
            folha = arquivo.active
            folha['A1'] = "Email:"
            folha['B1'] = "Senha:"
            arquivo.save("clientes.xlsx")
class Cliente():
    def Inserir_Cliente():
        arquivo = openpyxl.load_workbook('clientes.xlsx')
        folha = arquivo.active
        pessoa = {
            "email":request.form.get("email"),
            "senha":request.form.get("senha")
        }
        planilha = pd.read_excel('clientes.xlsx')
        resultado = planilha.loc[planilha['Email:'] == pessoa['email'], 'Email:']
        print(resultado)
# Verificar se algum email foi encontrado
        if not resultado.empty:
            print(f'O email  é: {resultado.iloc[0]}')
            return 1
        else:
            print(f'Cliente não encontrado.')
            folha.cell(column=1, row=folha.max_row+1, value=pessoa['email'])
            folha.cell(column=2, row=folha.max_row, value=pessoa['senha'])
            arquivo.save(r"clientes.xlsx")
            print("dados salvos com sucesso")
            return 0