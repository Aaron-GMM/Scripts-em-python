import openpyxl
import pandas as pd
import re
import pathlib
from openpyxl import Workbook
from flask import request
import datetime


class PlanilhaExcel:
    def criar_planilha():
        arquivo = pathlib.Path('Clientes.xlsx')
        if not arquivo.exists():
            arquivo = Workbook()
            folha = arquivo.active
            folha['A1'] = "Nome"
            folha['B1'] = "Whastapp"
            folha['C1'] = "Email"
            folha['D1'] = "Autoriza_contato"
            folha['E1'] = "Categoria"
            folha['F1'] = "Data de Acesso"
            folha['G1'] = "Quantidade de Acesso"
            arquivo.save("Clientes.xlsx")


class Cliente:
    def validar_nome(nome):
        return all(caracter.isalpha() or caracter.isspace() for caracter in nome)

    def validar_telefone(telefone):
        return all(numero.isdigit() or numero.isspace() for numero in telefone)

    def validar_email(email):
        return re.match("^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$", email)

    def data_atual():
        data_atual = datetime.date.today()
        data_formatada = data_atual.strftime('%d/%m/%Y')
        return data_formatada

    def inserir_cliente():
        arquivo = openpyxl.load_workbook("Clientes.xlsx")
        pessoa = {
            "nome": request.form.get("nome"),
            "whatsapp": request.form.get("whatsapp"),
            "email": request.form.get("email"),
            "categoria": request.form.get("categoria"),
            "autoriza_contato": request.form.get("autorizar_contato")
        }
        folha = arquivo.active
        if all(pessoa.values()):
            if (
                Cliente.validar_nome(pessoa["nome"]) and
                Cliente.validar_telefone(pessoa["whatsapp"]) and
                Cliente.validar_email(pessoa["email"])
            ):
                print("OK")
                planilha = pd.read_excel("Clientes.xlsx")
                resultado = planilha.loc[planilha['Email']
                                         == pessoa["email"], 'Email']
                if not resultado.empty:
                    return 1  # Email existente
                else:
                    print("Nao existente")
                    date = Cliente.data_atual()
                    folha.cell(column=1, row=folha.max_row +
                               1, value=pessoa["nome"])
                    folha.cell(column=2, row=folha.max_row,
                               value=pessoa["whatsapp"])
                    folha.cell(column=3, row=folha.max_row,
                               value=pessoa["email"])
                    folha.cell(column=4, row=folha.max_row,
                               value=pessoa["autoriza_contato"])
                    folha.cell(column=5, row=folha.max_row,
                               value=pessoa["categoria"])
                    folha.cell(column=6, row=folha.max_row, value=date)
                    folha.cell(column=7, row=folha.max_row, value=1)
                    arquivo.save("Clientes.xlsx")
                    print("---------------Dados Salvos com Sucesso!-----------------")
                    return 0  # Sucesso
            else:
                print("Campo errado")
                return 3  # Campo errado
        else:
            print("Campo Vazio")
            return 2  # Campo vazio


class login():
    def get_login():
        email = request.form.get('email')

        if not email:
            return 1

        if not Cliente.validar_email(email):
            return 2
        
        planilha = pd.read_excel("Clientes.xlsx")
        resultado = planilha[planilha['Email'] == email]

        if resultado.empty:
            return 3
        else:
            arquivo = openpyxl.load_workbook("Clientes.xlsx")
            folha = arquivo.active
            linhas_correspondentes = resultado
            linhas_correspondentes['Quantidade de Acesso'] = pd.to_numeric(linhas_correspondentes['Quantidade de Acesso'], errors='coerce', downcast='integer')
            quantidade_acesso = linhas_correspondentes.iloc[0]['Quantidade de Acesso']
            quantidade_acesso += 1
            date = Cliente.data_atual()
            folha.cell(column=6, row=folha.max_row, value=date)
            folha.cell(column=7, row=folha.max_row, value=quantidade_acesso)
            arquivo.save("Clientes.xlsx")
            print("---------------Dados Salvos com Sucesso!-----------------")
            return 0