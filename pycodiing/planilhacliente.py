from flask import request
import openpyxl
from openpyxl import Workbook
import pandas as pd
import pathlib
import re 

class Planilha():
    def criar_planilha():
     
     arquivo = pathlib.Path('Clientes.xlsx')

     if arquivo.exists():
         pass
     else:
        arquivo = Workbook()
        folha = arquivo.active
        folha['A1'] = "Nome"
        folha['B1'] = "Whastapp"
        folha['C1'] = "Email"
        folha['D1'] = "Autoriza_contato"
        folha['E1'] = "categoria"
        arquivo.save("Clientes.xlsx")

class Cliente():
   def inserir_client():
        arquivo = openpyxl.load_workbook("Clientes.xlsx")
        pessoa = {
                "nome":request.form.get("nome"),
                "whatsapp": request.form.get("whatsapp"),
                "email": request.form.get("email"),
                "categoria":  request.form.get("categoria"),
                "autoriza_contato": request.form.get("autorizar_contato")
        }
        folha = arquivo.active
        print("\n",pessoa["nome"], pessoa["email"], pessoa["whatsapp"], pessoa['categoria'], pessoa["autoriza_contato"],"\n")
        
        def validar_nome(pessoa):
            for caracter in pessoa:
                if not (caracter.isalpha() or caracter.isspace()):
                 return False
            return True
        
        def validar_telefone(pessoa):
            for numeber in pessoa:
                if not (numeber.isdigit() or numeber.isspace()):
                    return False
            
            return True
        def validar_email(pessoa):
             if re.match("^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$",pessoa):
               return 1
             else:
                return 0
        def validacao(pessoa):
           if "nome" in pessoa and validar_nome(pessoa['nome']):
              print("Nome valido")
              
           else:
              print("Nome invalido")
              return False
           if "whatsapp" in pessoa and validar_telefone(pessoa['whatsapp']):
              print("Telefone valido")
              
           else:
              print("Telefone Invalido")
              return False
              
           if "email" in pessoa and validar_email(pessoa['email']):
              print("Email valido")
           else:
              print("Email Invalido")
              return False

        if pessoa['nome'] != "" and pessoa["email"] != "" and pessoa['categoria'] != "" and pessoa["autoriza_contato"] != "":
           if validacao(pessoa)==False:
                print("campo errado")
                return 3
           else: 
            print("OK")
            planilha = pd.read_excel("Clientes.xlsx")
            resultado = planilha.loc[planilha['Email'] == pessoa["email"], 'Email']
            

            if not resultado.empty:
                print(f"Email existente")
                return 1
            else:
                print("Nao existente")
                folha.cell(column=1, row=folha.max_row+1, value=pessoa["nome"])
                folha.cell(column=2, row=folha.max_row,   value=pessoa["whatsapp"])
                folha.cell(column=3, row=folha.max_row,   value=pessoa["email"])
                folha.cell(column=4, row=folha.max_row,   value=pessoa["autoriza_contato"])
                folha.cell(column=5, row=folha.max_row,   value=pessoa["categoria"])
                arquivo.save(r"Clientes.xlsx")
                print("---------------Dados Salvos com Sucesso!-----------------")
                return 0
        else:
         print("Campo Vazio")
         return 2

